# excel/excel_exporter.py

import openpyxl
from openpyxl.styles import PatternFill, Font
import os
from datetime import datetime
from dotenv import load_dotenv
import subprocess
import platform

load_dotenv()

BASE_EXCEL = os.getenv("EXCEL_BASE_PATH", "base.xlsx")
OUTPUT_DIR = os.getenv("EXCEL_OUTPUT_DIR", "output")


def export_proforma_to_excel(model):
    if not os.path.exists(BASE_EXCEL):
        raise FileNotFoundError(f"No se encuentra el Excel base: {BASE_EXCEL}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)




    wb = openpyxl.load_workbook(BASE_EXCEL)
    ws = wb.active

    # ─────────────────────────────
    # Datos generales (opcionales)
    # ─────────────────────────────

    client_name = safe_get(model, "client_name")
    address = safe_get(model, "address")
    city = safe_get(model, "city")
    country = safe_get(model, "country", "España")
    contact = safe_get(model, "contact")

    cif = safe_get(model, "cif")
    postal_code = safe_get(model, "postal_code")
    province = safe_get(model, "province")
    phone = safe_get(model, "phone")
    email = safe_get(model, "email")

    area_m2 = safe_get(model, "area_m2")
    discount = safe_get(model, "discount_percent")
    shipping = safe_get(model, "shipping_cost")

    created_at = safe_get(model, "created_at", datetime.now().strftime("%d/%m/%Y"))
    proforma_id = safe_get(model, "proforma_id", None)

    proforma_number = format_proforma_number(proforma_id)


    # ─────────────────────────────
    # Posición inicial (B19)
    # ─────────────────────────────
    start_row = 19
    start_col = 2  # columna B
    current_row = start_row

    # ─────────────────────────────
    # Cliente
    # ─────────────────────────────
    ws["C9"] = client_name
    ws["C10"] = address
    ws["C11"] = city
    ws["C12"] = country
    ws["C13"] = contact

    ws["E9"] = cif
    ws["E10"] = postal_code
    ws["E11"] = province
    ws["E12"] = phone
    ws["E13"] = email

    # ─────────────────────────────
    # Proforma
    # ─────────────────────────────
    ws["D5"] = f"Proforma: {proforma_number}"
    ws["D6"] = f"Fecha: {created_at}"

    # ─────────────────────────────
    # Totales
    # ─────────────────────────────
    if discount != "":
        ws["D43"] = discount

    if shipping != "":
        ws["E45"] = shipping


    # ─────────────────────────────
    # Estilos
    # ─────────────────────────────
    default_font = Font(name="Calibri", size=12, bold=True)

    title_fill = PatternFill(
        start_color="0000FF",
        end_color="0000FF",
        fill_type="solid"
    )
    title_font = Font(
        name="Calibri",
        size=12,
        color="FFFFFF",
        bold=True
    )

    # ─────────────────────────────
    # Render filas
    # ─────────────────────────────
    for row in model.rows:

        # =========================
        # TITLE
        # =========================
        if row.type == "TITLE":
            ws.merge_cells(
                start_row=current_row,
                start_column=start_col,
                end_row=current_row,
                end_column=start_col + 4
            )

            cell = ws.cell(
                row=current_row,
                column=start_col,
                value=row.col_0
            )
            cell.fill = title_fill
            cell.font = title_font

            current_row += 1

        # =========================
        # PRODUCT
        # =========================
        elif row.type == "PRODUCT":
            if float(row.col_2) > 0:
                # B → Kits
                ws.cell(
                    row=current_row,
                    column=start_col,
                    value=row.col_0
                ).font = default_font

                # C → Producto
                ws.cell(
                    row=current_row,
                    column=start_col + 1,
                    value=row.col_1
                ).font = default_font

                # D → Cantidad
                qty_cell = ws.cell(
                    row=current_row,
                    column=start_col + 2,
                    value=row.col_2
                )
                qty_cell.font = default_font

                # E → Precio €/Ud
                price_cell = ws.cell(
                    row=current_row,
                    column=start_col + 3,
                    value=row.col_3
                )
                price_cell.font = default_font

                # F → Total (fórmula Excel)
                total_cell = ws.cell(
                    row=current_row,
                    column=start_col + 4,
                    value=f"={qty_cell.coordinate}*{price_cell.coordinate}"
                )
                total_cell.font = default_font

                current_row += 1

        # =========================
        # INFO (comentarios)
        # =========================
        elif row.type == "INFO":
            ws.merge_cells(
                start_row=current_row,
                start_column=start_col,
                end_row=current_row,
                end_column=start_col + 4
            )

            cell = ws.cell(
                row=current_row,
                column=start_col,
                value=row.col_0
            )
            cell.font = default_font

            current_row += 1

        # =========================
        # EMPTY
        # =========================
        elif row.type == "EMPTY":
            current_row += 1

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = client_name.replace(" ", "_") if client_name else "cliente"
    safe_area = f"{area_m2}m" if area_m2 else ""

    output_path = os.path.join(
        OUTPUT_DIR,
        f"{proforma_number}_{safe_name}_{safe_area}.xlsx"
    )

    wb.save(output_path)

    # ─────────────────────────────
    # Abrir automáticamente
    # ─────────────────────────────
    try:
        if platform.system() == "Windows":
            os.startfile(output_path)
        elif platform.system() == "Darwin":
            subprocess.call(["open", output_path])
        else:
            subprocess.call(["xdg-open", output_path])
    except Exception as e:
        print(f"No se pudo abrir automáticamente el archivo: {e}")

    return output_path


def format_proforma_number(proforma_id: int | None) -> str:
    if proforma_id is None:
        return datetime.now().strftime("%Y%m%d")
    return str(proforma_id).zfill(6)

def safe_get(obj, attr, default=""):
    return getattr(obj, attr, default) or default
